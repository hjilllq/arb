import csv
from pathlib import Path

from openpyxl import load_workbook

import numpy as np

from data_exporter import (
    export_trades_to_csv,
    export_trades_to_excel,
    export_trades_to_npz,
)


SAMPLE = [
    {"pair": "BTCUSDT", "price": 100.0, "qty": 1, "timestamp": "2024-01-01"}
]


def test_export_csv(tmp_path: Path) -> None:
    file_path = tmp_path / "trades.csv"
    assert export_trades_to_csv(SAMPLE, file_path)
    with file_path.open(newline="", encoding="utf-8") as fh:
        rows = list(csv.DictReader(fh))
    assert rows[0]["pair"] == "BTCUSDT"


def test_export_excel(tmp_path: Path) -> None:
    file_path = tmp_path / "trades.xlsx"
    assert export_trades_to_excel(SAMPLE, file_path)
    wb = load_workbook(file_path)
    ws = wb.active
    assert ws.cell(row=2, column=1).value == "BTCUSDT"


def test_export_npz(tmp_path: Path) -> None:
    file_path = tmp_path / "trades.npz"
    assert export_trades_to_npz(SAMPLE, file_path)
    data = np.load(file_path)
    assert data["pair"][0] == "BTCUSDT"
